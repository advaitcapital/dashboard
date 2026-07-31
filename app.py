#!/usr/bin/env python3
"""
LIVE Market Dashboard (Streamlit) - 9-section management layout
===============================================================
Auto-refreshes. Every number comes from a real source and is labelled.
Items with a free feed update automatically; items without a free source
show a blank "-" and are never invented. Add a free FRED key in the sidebar
to unlock the US/EU/India macro rows.

RUN:  python3 -m streamlit run app.py
"""

import datetime as dt
import html as _html
import io
import json
import os
import re

import pandas as pd
import requests
import streamlit as st

try:
    from streamlit_autorefresh import st_autorefresh
    HAS_AUTOREFRESH = True
except Exception:
    HAS_AUTOREFRESH = False

try:
    import yfinance as yf
    HAS_YF = True
except Exception:
    HAS_YF = False

try:
    import altair as _alt
    HAS_ALT = True
except Exception:
    HAS_ALT = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    HAS_XLSX = True
except Exception:
    HAS_XLSX = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, Image)
    HAS_PDF = True
except Exception:
    HAS_PDF = False

H = {"User-Agent": "Mozilla/5.0 (personal market dashboard)"}
MANUAL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "manual_data.json")
# Local timezone label, correct on any machine/OS (e.g. IST, EST, GMT)
TZ = dt.datetime.now().astimezone().tzname() or "local"



def load_manual():
    try:
        with open(MANUAL_FILE, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_manual(d):
    """Merge the given keys into manual_data.json (keeps other saved fields)."""
    try:
        cur = load_manual()
        cur.update(d)
        with open(MANUAL_FILE, "w", encoding="utf-8") as f:
            json.dump(cur, f, indent=2)
        return True
    except Exception:
        return False


manual = load_manual()
FRED_API_KEY = manual.get("fred_key", "").strip()

import report_pdf as _report_pdf   # shared PDF builder
DASHBOARD_URL = "https://market-dashboard-7hsncprqzkfytpx7a8ishb.streamlit.app"

# Latest published India macro that has NO free real-time API (released only in
# MoSPI/PIB press releases). These are the defaults shown until you override them
# in the sidebar ("India macro" box). Update when a new release comes out:
#   GDP  -> quarterly by MoSPI (NSO), ~last business day of the 2nd month after
#   WPI  -> monthly by the Office of the Economic Adviser, ~14th of next month
# Latest published India figures. No free live API exists for these (MoSPI /
# DPIIT release them by press note; FRED's OECD-MEI India CPI series stopped at
# Mar 2025). Each entry is (latest, period, previous) so the table can show the
# real month-on-month move. Update via the sidebar when a new print lands.
LATEST_INDIA_GDP = ("7.70", "FY26", "6.60")      # FY26 actual; FY27 RBI forecast 6.6%
LATEST_INDIA_WPI = ("9.87", "Jun 2026", "9.68")  # new 2022-23 base series
LATEST_INDIA_CPI = ("4.38", "Jun 2026", "3.93")  # released 13 Jul 2026
# India G-Sec yields — no free API; RBI/FBIL publishes daily but behind login.
# Source: TradingEconomics / RBI FBIL benchmark. Update via sidebar.
LATEST_INDIA_1Y = 5.92    # 52-week T-bill yield
LATEST_INDIA_2Y = 6.25    # 2-year G-Sec yield

# ---------------------------------------------------------------------------
# TICKERS (Yahoo Finance)
# ---------------------------------------------------------------------------
EQUITIES = [
    ("India", "Nifty 50", "^NSEI"), ("India", "Sensex", "^BSESN"),
    ("US", "S&P 500", "^GSPC"), ("US", "Nasdaq", "^IXIC"),
    ("US", "Dow Jones", "^DJI"), ("Germany", "DAX", "^GDAXI"),
    ("UK", "FTSE 100", "^FTSE"), ("China", "Shanghai Composite", "000001.SS"),
    ("Hong Kong", "Hang Seng", "^HSI"), ("Japan", "Nikkei 225", "^N225"),
]
FX = [("USD/INR", "INR=X"), ("EUR/INR", "EURINR=X"), ("GBP/INR", "GBPINR=X"),
      ("JPY/INR", "JPYINR=X"), ("USD/CNY", "CNY=X")]
_OZ_G = 31.1035       # grams per troy ounce
_LB_KG = 0.453592     # kg per pound
# (label, yahoo ticker, factor to convert the USD native-unit price to the unit)
# (label, yahoo ticker, INR factor, intl_unit_label, intl_factor)
# intl_factor converts yahoo's native unit to the displayed intl unit.
COMMODITIES = [
    ("Gold 24K",      "GC=F",  10 / _OZ_G, "USD/oz",    1.0),
    ("Silver",        "SI=F",  1000/_OZ_G,  "USD/oz",    1.0),
    ("Platinum",      "PL=F",  10 / _OZ_G, "USD/oz",    1.0),
    ("Copper",        "HG=F",  1/_LB_KG,    "USD/lb",    1.0),
    ("Aluminum",      "ALI=F", 1.0,         "USD/t",     1.0),
    ("Palladium",     "PA=F",  10 / _OZ_G, "USD/oz",    1.0),
    ("Uranium (Sprott)","SRUUF",  1/_LB_KG,    "USD/lb",    1.0),
    ("Brent Crude",   "BZ=F",  1.0,         "USD/bbl",   1.0),
    ("WTI Crude",     "CL=F",  1.0,         "USD/bbl",   1.0),
    ("Natural Gas",   "NG=F",  1.0,         "USD/MMBtu", 1.0),
]
VOL = [("India VIX", "^INDIAVIX"), ("CBOE VIX", "^VIX")]
REITS = [("Embassy REIT", "EMBASSY.BO"), ("Mindspace REIT", "MINDSPACE.BO"),
         ("Brookfield REIT", "BIRET.BO"), ("Nexus Select REIT", "NXST.BO")]
INVITS = [("IndiGrid InvIT", "INDIGRID.BO"), ("Powergrid InvIT", "PGINVIT.BO"),
          ("IRB InvIT", "IRBINVIT.BO")]
CRYPTO = [("Bitcoin", "BTC-USD"), ("Ethereum", "ETH-USD")]



# ---------------------------------------------------------------------------
# FETCHERS
# ---------------------------------------------------------------------------
@st.cache_data(ttl=300)
def yahoo(tickers):
    out = {}
    if not HAS_YF:
        return out
    for item in tickers:
        tk = item[-1]
        try:
            h = yf.Ticker(tk).history(period="6y")["Close"].dropna()
            if len(h) < 2:
                h = yf.Ticker(tk).history(period="5d")["Close"].dropna()
            if len(h) >= 2:
                p, prev = float(h.iloc[-1]), float(h.iloc[-2])
                rec = {"price": p, "d1": (p/prev-1)*100, "base": {"d1": prev}}
                # YTD: first close of the current calendar year
                this_year = h[h.index.year == h.index[-1].year]
                ys = float(this_year.iloc[0]) if len(this_year) else float(h.iloc[0])
                rec["ytd"] = (p/ys-1)*100
                rec["base"]["ytd"] = ys
                # 1M / 3M / 6M / 1Y / 5Y: close nearest to N days ago
                last_date = h.index[-1]
                for label, days in (("m1", 30), ("m3", 91), ("m6", 182),
                                    ("y1", 365), ("y5", 1826)):
                    cutoff = last_date - pd.Timedelta(days=days)
                    past = h[h.index <= cutoff]
                    base = float(past.iloc[-1]) if len(past) else None
                    rec[label] = (p/base-1)*100 if base else None
                    rec["base"][label] = base
                out[tk] = rec
        except Exception:
            pass
    return out


@st.cache_data(ttl=21600)  # 6h — distributions are announced quarterly
def ttm_payout(tk):
    """Sum of the ACTUAL distributions Yahoo recorded in the trailing 12 months.
    Real data imported from Yahoo, not an assumption. Returns rupees/unit, or
    None if Yahoo has no distribution history for this ticker."""
    if not HAS_YF:
        return None
    try:
        s = yf.Ticker(tk).dividends            # pandas Series, indexed by pay date
        if s is None or len(s) == 0:
            return None
        now = pd.Timestamp.now(tz=s.index.tz) if s.index.tz is not None \
            else pd.Timestamp.now()
        total = float(s[s.index >= now - pd.Timedelta(days=365)].sum())
        return total if total > 0 else None
    except Exception:
        return None


@st.cache_data(ttl=300)
@st.cache_data(ttl=3600)  # 1h — treasury.gov updates once daily; caching also
def treasury():           # avoids hammering (and being blocked by) the gov site
    yr = dt.date.today().year
    url = (f"https://home.treasury.gov/resource-center/data-chart-center/"
           f"interest-rates/daily-treasury-rates.csv/{yr}/all"
           f"?type=daily_treasury_yield_curve&field_tdr_date_value={yr}"
           f"&page&_format=csv")
    r = requests.get(url, headers=H, timeout=20); r.raise_for_status()
    df = pd.read_csv(io.StringIO(r.text))
    df["Date"] = pd.to_datetime(df["Date"]); df = df.sort_values("Date")
    a, b = df.iloc[-1], df.iloc[-2]
    g = lambda c: (float(a[c]), float(a[c]-b[c])) if c in df.columns else (None, None)
    return {"1Y": g("1 Yr"), "2Y": g("2 Yr"), "10Y": g("10 Yr"),
            "date": a["Date"].date().isoformat()}


@st.cache_data(ttl=3600)  # 1h — JGB curve updates once daily
def japan_jgb():
    """Japan JGB yields from the Ministry of Finance daily CSV (free, plain HTTP,
    works on Streamlit Cloud — no browser needed). The CSV has a full curve:
    Date,1Y,2Y,3Y,...,10Y,...,40Y. Returns
    {'1Y':float,'2Y':float,'10Y':float,'10Y_chg':float(pct-points)} or {}."""
    url = ("https://www.mof.go.jp/english/policy/jgbs/reference/"
           "interest_rate/jgbcme.csv")
    try:
        r = requests.get(url, headers=H, timeout=20)
        if r.status_code != 200 or not r.text.strip():
            return {}
        lines = [ln for ln in r.text.splitlines() if ln.strip()]
        hdr_i = next((i for i, l in enumerate(lines)
                      if l.split(",")[0].strip() == "Date"), None)
        if hdr_i is None:
            return {}
        header = [h.strip() for h in lines[hdr_i].split(",")]
        idx = {name: j for j, name in enumerate(header)}
        # Keep only real data rows (first cell is a date like 2026/5/18). The CSV
        # ends with footer junk (an all-commas line and a "clear your cache"
        # message) — those must be skipped or the latest row parses as empty.
        rows = [l.split(",") for l in lines[hdr_i + 1:]
                if re.match(r"^\s*\d{4}/\d{1,2}/\d{1,2}", l.split(",")[0])]
        if not rows:
            return {}

        def val(row, col):
            try:
                return float(row[idx[col]])
            except Exception:
                return None

        last = rows[-1]
        out = {"1Y": val(last, "1Y"), "2Y": val(last, "2Y"),
               "10Y": val(last, "10Y")}
        if len(rows) >= 2 and out["10Y"] is not None:
            p10 = val(rows[-2], "10Y")
            if p10 is not None:
                out["10Y_chg"] = round(out["10Y"] - p10, 3)
        return out
    except Exception:
        return {}


@st.cache_data(ttl=3600)  # 1h — euro-area curve updates daily ~noon CET
def ecb_yield():
    """Euro-area AAA government bond yields (the German-bund benchmark curve)
    from the ECB Data Portal SDMX REST API — free, plain HTTP, works on cloud.
    Returns {'1Y':float,'2Y':float,'10Y':float} or {}. Used for 'Germany'."""
    base = ("https://data-api.ecb.europa.eu/service/data/YC/"
            "B.U2.EUR.4F.G_N_A.SV_C_YM.SR_{}?lastNObservations=1&format=csvdata")
    out = {}
    for tenor in ("1Y", "2Y", "10Y"):
        try:
            r = requests.get(base.format(tenor), headers=H, timeout=20)
            if r.status_code != 200 or not r.text.strip():
                continue
            lines = [ln for ln in r.text.splitlines() if ln.strip()]
            if len(lines) < 2:
                continue
            cols = lines[0].split(",")
            vi = cols.index("OBS_VALUE") if "OBS_VALUE" in cols else -1
            last = lines[-1].split(",")
            if vi >= 0 and vi < len(last):
                out[tenor] = round(float(last[vi]), 2)
        except Exception:
            pass
    return out


@st.cache_data(ttl=3600)  # 1h — BoE publishes daily
def boe_yields():
    """UK gilt yields from the Bank of England CSV API (free, plain HTTP, works
    on Streamlit Cloud). IUDSNPY = short-term nominal par yield (~1Y),
    IUDMNPY = medium-term nominal par yield (~2-5Y).
    Returns {'1Y': float, '2Y': float} or {}."""
    today = dt.date.today()
    dfrom = (today - dt.timedelta(days=30)).strftime("%d/%b/%Y")
    url = ("https://www.bankofengland.co.uk/boeapps/database/"
           "_iadb-fromshowcolumns.asp?csv.x=yes"
           f"&SeriesCodes=IUDSNPY,IUDMNPY&UsingCodes=Y&CSVF=TN&Datefrom={dfrom}")
    try:
        r = requests.get(url, headers=H, timeout=20)
        if r.status_code != 200 or not r.text.strip():
            return {}
        lines = [l for l in r.text.splitlines() if l.strip()]
        if len(lines) < 2:
            return {}
        hdr = [h.strip().upper() for h in lines[0].split(",")]
        out = {}
        last = lines[-1].split(",")
        for i, col in enumerate(hdr):
            if "IUDSNPY" in col and i < len(last):
                try:
                    out["1Y"] = round(float(last[i].strip()), 2)
                except Exception:
                    pass
            if "IUDMNPY" in col and i < len(last):
                try:
                    out["2Y"] = round(float(last[i].strip()), 2)
                except Exception:
                    pass
        return out
    except Exception:
        return {}


@st.cache_data(ttl=3600)
def india_short_yields():
    """India 1Y/2Y G-Sec yields. No free API exists — FRED's OECD-MEI India
    series were discontinued. We use the latest published figures (RBI/FBIL),
    overridable in the sidebar. Returns {'1Y': float, '2Y': float} or {}."""
    out = {}
    # Try sidebar manual override first
    y1_m = str(manual.get("india_1y", "")).strip()
    y2_m = str(manual.get("india_2y", "")).strip()
    try:
        out["1Y"] = float(y1_m) if y1_m else LATEST_INDIA_1Y
    except Exception:
        out["1Y"] = LATEST_INDIA_1Y
    try:
        out["2Y"] = float(y2_m) if y2_m else LATEST_INDIA_2Y
    except Exception:
        out["2Y"] = LATEST_INDIA_2Y
    return out


@st.cache_data(ttl=120)
def crypto():
    url = ("https://api.coingecko.com/api/v3/simple/price?ids=bitcoin,ethereum"
           "&vs_currencies=usd&include_24hr_change=true")
    r = requests.get(url, headers=H, timeout=20); r.raise_for_status()
    d = r.json()
    return {"Bitcoin": (d["bitcoin"]["usd"], d["bitcoin"]["usd_24h_change"]),
            "Ethereum": (d["ethereum"]["usd"], d["ethereum"]["usd_24h_change"])}


@st.cache_data(ttl=1800)  # 30 min — Indian gold rate updates a few times a day
def india_gold():
    """LIVE Indian retail 24K gold rate in INR per gram, scraped from Goodreturns
    (server-rendered, no JS). This is the actual Indian quoted rate — it INCLUDES
    the duty/premium that the COMEX-spot conversion omits, so it matches what
    jewellers and other Indian sites show. Returns {'24k': float} per gram, or {}."""
    srcs = ["https://www.goodreturns.in/gold-rates/",
            "https://www.goodreturns.in/gold-rates/mumbai.html",
            "https://www.goodreturns.in/gold-rates/delhi.html"]
    for url in srcs:
        try:
            r = requests.get(url, headers=_BROWSER_H, timeout=15)
            if r.status_code != 200:
                continue
            txt = re.sub(r"\s+", " ",
                         _html.unescape(_TAGS.sub(" ", _STRIP.sub(" ", r.text))))
            m24 = re.search(r"([\d,]{3,})\s*per\s*gram\s*for\s*24", txt, re.I)
            g24 = float(m24.group(1).replace(",", "")) if m24 else None
            # sanity: Indian 24K/gram is in the thousands (₹), reject junk
            if g24 and 1000 <= g24 <= 100000:
                return {"24k": g24}
        except Exception:
            pass
    return {}


@st.cache_data(ttl=600)
def fred(series, key):
    if not key:
        return None
    try:
        url = (f"https://api.stlouisfed.org/fred/series/observations?series_id="
               f"{series}&api_key={key}&file_type=json&sort_order=desc&limit=1")
        v = requests.get(url, timeout=20).json()["observations"][0]["value"]
        return float(v)          # FRED sends "." for missing -> ValueError -> None
    except Exception:
        return None


@st.cache_data(ttl=21600)  # 6h — monthly macro series
def fred_latest(series, key):
    """Latest (value, 'Mon YYYY') for a FRED series, skipping missing '.' obs.
    Used for current monthly macro figures. Returns (float, str) or (None, None)."""
    if not key:
        return None, None
    try:
        url = (f"https://api.stlouisfed.org/fred/series/observations?series_id="
               f"{series}&api_key={key}&file_type=json&sort_order=desc&limit=12")
        obs = requests.get(url, timeout=20).json().get("observations", [])
        for o in obs:
            v = o.get("value")
            if v not in (".", "", None):
                d = o.get("date", "")
                try:
                    label = dt.datetime.strptime(d, "%Y-%m-%d").strftime("%b %Y")
                except Exception:
                    label = d
                return float(v), label
        return None, None
    except Exception:
        return None, None


@st.cache_data(ttl=21600)  # 6h — historical curve barely changes intraday
def fred_history(series, key, years=5):
    """Full observation history for a FRED series over the last `years` years.
    Returns a date-indexed pandas Series of floats (missing '.' values dropped),
    or None. Daily series (e.g. DGS10) come back daily; OECD cross-country
    series come back monthly. Used by the 5-year 10Y yield chart."""
    if not key:
        return None
    try:
        start = (dt.date.today() - dt.timedelta(days=365 * years + 10)).isoformat()
        url = (f"https://api.stlouisfed.org/fred/series/observations?series_id="
               f"{series}&api_key={key}&file_type=json&sort_order=asc"
               f"&observation_start={start}")
        obs = requests.get(url, timeout=25).json().get("observations", [])
        pairs = [(o["date"], float(o["value"])) for o in obs
                 if o.get("value") not in (".", "", None)]
        if not pairs:
            return None
        s = pd.Series({pd.to_datetime(d): v for d, v in pairs}).sort_index()
        return s if len(s) else None
    except Exception:
        return None


@st.cache_data(ttl=86400)
def policy_rate_changes(series, key):
    """Policy-rate level + change (in bps) over 1M/3M/6M/1Y/5Y from a FRED
    series. Returns (latest, {'m1':bps, 'm3':bps, ...}) or (None, {})."""
    s = fred_history(series, key, years=5)
    if s is None or not len(s):
        return None, {}
    latest = float(s.iloc[-1])
    now = s.index[-1]
    out = {}
    for lbl, days in (("m1", 30), ("m3", 91), ("m6", 182),
                      ("y1", 365), ("y5", 365 * 5)):
        cutoff = now - pd.Timedelta(days=days)
        prior = s[s.index <= cutoff]
        if len(prior):
            out[lbl] = round((latest - float(prior.iloc[-1])) * 100, 0)  # bps
    return latest, out


@st.cache_data(ttl=86400)
def macro_yoy(series, key):
    """Latest level + year-on-year % change + observation date for a FRED series.
    Returns (level, yoy_pct, date_str) — any part may be None."""
    s = fred_history(series, key, years=6)
    if s is None or not len(s):
        return None, None, None
    level = float(s.iloc[-1])
    last = s.index[-1]
    date_str = last.strftime("%b %Y")
    prior = s[s.index <= last - pd.Timedelta(days=365)]
    yoy = ((level / float(prior.iloc[-1]) - 1) * 100
           if len(prior) and float(prior.iloc[-1]) else None)
    return level, yoy, date_str


def _macro_cells(series, key, unit="", as_change=False):
    """(value cell, change cell) for the macro table.
    as_change=True means the series is already a rate (e.g. unemployment %),
    so the 'change' column shows the move in percentage points, not a % of a %."""
    level, yoy, date = macro_yoy(series, key)
    if level is None:
        return gap("add FRED key"), '<span class="muted">NA</span>'
    val = f"{level:,.2f}{unit}"
    if date:
        val += f' <span class="muted">({date})</span>'
    if yoy is None:
        return val, '<span class="muted">NA</span>'
    if as_change:
        # percentage-point move over 12 months
        s = fred_history(series, key, years=6)
        prior = s[s.index <= s.index[-1] - pd.Timedelta(days=365)]
        pp = level - float(prior.iloc[-1]) if len(prior) else None
        return val, (chg(pp, bps=False) if pp is not None
                     else '<span class="muted">NA</span>')
    return val, chg(yoy)


@st.cache_data(ttl=86400)
def bis_policy(cc):
    """Central bank policy rate from BIS (free, weekly). cc = IN, GB, XM, US.
    Best-effort across a few endpoint shapes; returns '5.25%' style or None."""
    urls = [
        f"https://stats.bis.org/api/v1/data/WS_CBPOL/D.{cc}/all?lastNObservations=1&format=csv",
        f"https://stats.bis.org/api/v1/data/WS_CBPOL/M.{cc}/all?lastNObservations=1&format=csv",
        f"https://stats.bis.org/api/v1/data/BIS,WS_CBPOL,1.0/D.{cc}?lastNObservations=1&format=csv",
        f"https://stats.bis.org/api/v1/data/BIS,WS_CBPOL,1.0/M.{cc}?lastNObservations=1&format=csv",
    ]
    for url in urls:
        try:
            r = requests.get(url, headers=H, timeout=20)
            if r.status_code == 200 and r.text.strip():
                df = pd.read_csv(io.StringIO(r.text))
                vcol = [c for c in df.columns if c.upper() == "OBS_VALUE"]
                if vcol:
                    val = pd.to_numeric(df[vcol[0]], errors="coerce").dropna()
                    if len(val):
                        return f"{float(val.iloc[-1]):.2f}%"
        except Exception:
            pass
    return None


@st.cache_data(ttl=3600)
def worldbank(indicator):
    """India macro from the World Bank API (free, no key). Returns (value, year).
    Scans the last 10 years newest-first and retries once, so a transient miss
    doesn't get stuck."""
    for _ in range(2):
        try:
            url = (f"https://api.worldbank.org/v2/country/IND/indicator/{indicator}"
                   f"?format=json&mrv=10")
            r = requests.get(url, headers=H, timeout=25); r.raise_for_status()
            data = r.json()
            if isinstance(data, list) and len(data) > 1 and data[1]:
                for obs in data[1]:  # newest first
                    if obs.get("value") is not None:
                        return float(obs["value"]), obs.get("date")
        except Exception:
            pass
    return None


# ---------------------------------------------------------------------------
# BANK FD SCRAPER (best-effort; only banks that serve rates in raw HTML)
# ---------------------------------------------------------------------------
_BROWSER_H = {
    "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/124.0 Safari/537.36"),
    "Accept-Language": "en-US,en;q=0.9",
}
_STRIP = re.compile(r"<script.*?</script>|<style.*?</style>", re.S | re.I)
_TAGS = re.compile(r"<[^>]+>")
# Every "<n> <unit> ... <rate>%" pair on the page.
_TENOR_RATE = re.compile(
    r"(\d{1,4})\s*(day|days|month|months|year|years|yr|yrs)\b[^%]{0,45}?"
    r"([3-9]\.\d{1,2})\s*%", re.I)
_UNIT_DAYS = {"day": 1, "days": 1, "month": 30, "months": 30,
              "year": 365, "years": 365, "yr": 365, "yrs": 365}


@st.cache_data(ttl=21600)  # 6h — FD rates change only a few times a year
def bank_fd(url):
    """Scrape a bank's ~1-year FD rate from its public page via a simple fetch.
    Works for banks that serve rates in raw HTML (HDFC, Axis)."""
    try:
        r = requests.get(url, headers=_BROWSER_H, timeout=12)
        if r.status_code != 200:
            return None
        text = _html.unescape(_TAGS.sub(" ", _STRIP.sub(" ", r.text)))
        return _pick_1y(re.sub(r"\s+", " ", text))
    except Exception:
        return None


def _pick_1y(text):
    """From visible text, collect every (tenor->days, rate) pair and return the
    rate for the bucket nearest 365 days. Shared by the simple fetch and the
    headless browser. Returns {rate, days, cands} or None."""
    cands = []
    for m in _TENOR_RATE.finditer(text):
        n, unit, rate = int(m.group(1)), m.group(2).lower(), float(m.group(3))
        days = n * _UNIT_DAYS[unit]
        if 2.0 <= rate <= 9.5 and 7 <= days <= 3700:
            cands.append((days, rate))
    if not cands:
        return None
    best = min(cands, key=lambda c: (abs(c[0] - 365), c[0]))
    seen, sample = set(), []
    for d, rt in cands:
        if (d, rt) not in seen:
            seen.add((d, rt)); sample.append((d, rt))
        if len(sample) >= 12:
            break
    return {"rate": best[1], "days": best[0], "cands": sample}


@st.cache_data(ttl=21600)  # 6h — heavy call, so cache hard
def bank_fd_browser(url):
    """Scrape a JS-rendered bank page (SBI, ICICI, Kotak) using a real Chrome
    via Playwright: open the page, let the rate JS run, then parse the text.
    Returns {rate, days, cands} or None. Needs: pip install playwright;
    python -m playwright install chromium."""
    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        return None
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page(user_agent=_BROWSER_H["User-Agent"])
            page.goto(url, timeout=45000, wait_until="domcontentloaded")
            page.wait_for_timeout(6000)        # let the rate table load
            body = page.inner_text("body")
            browser.close()
        return _pick_1y(re.sub(r"\s+", " ", body))
    except Exception:
        return None


# ---------------------------------------------------------------------------
# MULTI-TENOR FD SCRAPER
# ---------------------------------------------------------------------------
_FD_TARGETS = {"1M": 30, "3M": 91, "6M": 182, "1Y": 365, "3Y": 1095}
_FD_TOL     = {"1M": 25, "3M": 50, "6M": 75,  "1Y": 120, "3Y": 250}

# Wider regex: rate range 2.x–9.x, longer gap so wide tables still match.
_TENOR_RATE_WIDE = re.compile(
    r"(\d{1,4})\s*(day|days|month|months|year|years|yr|yrs|mo|mth|mths)\b"
    r"[^%]{0,80}?([2-9]\.\d{1,2})\s*%", re.I)
_UNIT_DAYS_W = {"day": 1, "days": 1, "month": 30, "months": 30,
                "mo": 30, "mth": 30, "mths": 30,
                "year": 365, "years": 365, "yr": 365, "yrs": 365}


def _pick_all_tenors(text):
    """Collect every (tenor->days, rate) pair from page text and bucket them
    into the 5 target tenors. The FIRST rate after a tenor is taken (aggregator
    tables list general-public before senior-citizen), so we read the general
    public rate. Returns {tenor_label: 'X.XX'}."""
    cands = []
    for m in _TENOR_RATE_WIDE.finditer(text):
        n = int(m.group(1))
        unit = m.group(2).lower()
        rate = float(m.group(3))
        days = n * _UNIT_DAYS_W.get(unit, 0)
        if 2.0 <= rate <= 9.75 and 7 <= days <= 4000:
            cands.append((days, rate))
    result = {}
    for label, want in _FD_TARGETS.items():
        tol = _FD_TOL[label]
        near = [c for c in cands if abs(c[0] - want) <= tol]
        if near:
            best = min(near, key=lambda c: abs(c[0] - want))
            result[label] = f"{best[1]:.2f}"
    return result


def _merge_tenors(dst, src):
    """Fill only the tenors dst is still missing — earlier (higher-priority)
    sources win, later sources just patch the gaps."""
    for k, v in src.items():
        dst.setdefault(k, v)
    return dst


def _scrape_http(url):
    """Plain HTTP fetch + parse. Thread-safe. Returns {tenor: rate} or {}."""
    if not url:
        return {}
    try:
        r = requests.get(url, headers=_BROWSER_H, timeout=15)
        if r.status_code == 200:
            txt = _html.unescape(_TAGS.sub(" ", _STRIP.sub(" ", r.text)))
            return _pick_all_tenors(re.sub(r"\s+", " ", txt))
    except Exception:
        pass
    return {}


def _scrape_browser(url):
    """Headless-browser fetch + parse. MAIN THREAD ONLY (Playwright sync API
    cannot run inside a worker thread). Returns {tenor: rate} or {}."""
    if not url:
        return {}
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page(user_agent=_BROWSER_H["User-Agent"])
            page.goto(url, timeout=45000, wait_until="domcontentloaded")
            page.wait_for_timeout(7000)          # let rate JS finish
            try:                                  # nudge lazy tables into view
                page.mouse.wheel(0, 4000)
                page.wait_for_timeout(1500)
            except Exception:
                pass
            body = page.inner_text("body")
            browser.close()
        return _pick_all_tenors(re.sub(r"\s+", " ", body))
    except Exception:
        return {}


# Per bank: (name, [aggregator URLs in priority order], direct bank URL).
# BankBazaar pages are server-rendered with full tenure tables (most reliable),
# so they go first; PolicyBazaar (verified "-fd-rates" slug) and Paisabazaar
# follow. Results are MERGED across all of them, so a tenor missing on one site
# is filled by another. 404s on a guessed URL are harmless (ignored).
_FD_INSTITUTIONS = [
    ("SBI", [
        "https://www.bankbazaar.com/fixed-deposit/sbi-fixed-deposit-rate.html",
        "https://www.policybazaar.com/fd-interest-rates/sbi-fd-rates/",
        "https://www.paisabazaar.com/fixed-deposit/sbi-fd-rates/",
    ], "https://sbi.co.in/web/interest-rates/deposit-rates/retail-domestic-term-deposits"),
    ("HDFC Bank", [
        "https://www.bankbazaar.com/fixed-deposit/hdfc-fixed-deposit-rate.html",
        "https://www.policybazaar.com/fd-interest-rates/hdfc-bank-fd-rates/",
        "https://www.paisabazaar.com/fixed-deposit/hdfc-bank-fd-rates/",
    ], "https://www.hdfcbank.com/personal/save-invest/deposits/fixed-deposit/fixed-deposit-interest-rates"),
    ("ICICI Bank", [
        "https://www.bankbazaar.com/fixed-deposit/icici-fixed-deposit-rate.html",
        "https://www.policybazaar.com/fd-interest-rates/icici-bank-fd-rates/",
        "https://www.paisabazaar.com/fixed-deposit/icici-bank-fd-rates/",
    ], "https://www.icicibank.com/personal-banking/deposits/fixed-deposit/fd-interest-rates"),
    ("Axis Bank", [
        "https://www.bankbazaar.com/fixed-deposit/axis-fixed-deposit-rate.html",
        "https://www.policybazaar.com/fd-interest-rates/axis-bank-fd-rates/",
        "https://www.paisabazaar.com/fixed-deposit/axis-bank-fd-rates/",
    ], "https://www.axisbank.com/fixed-deposit-interest-rate"),
    ("Kotak", [
        "https://www.bankbazaar.com/fixed-deposit/kotak-mahindra-bank-fixed-deposit-rate.html",
        "https://www.policybazaar.com/fd-interest-rates/kotak-mahindra-bank-fd-rates/",
        "https://www.paisabazaar.com/fixed-deposit/kotak-mahindra-bank-fd-rates/",
    ], "https://www.kotak.com/en/personal-banking/deposits/fixed-deposit/fixed-deposit-interest-rate.html"),
    ("AU SFB", [
        "https://www.policybazaar.com/fd-interest-rates/au-small-finance-bank-fd-rates/",
        "https://www.bankbazaar.com/fixed-deposit/au-small-finance-bank-fixed-deposit-rate.html",
        "https://www.paisabazaar.com/fixed-deposit/au-small-finance-bank-fd-rates/",
    ], "https://www.aubank.in/interest-rates/fixed-deposit-interest-rates"),
    ("Ujjivan SFB", [
        "https://www.policybazaar.com/fd-interest-rates/ujjivan-small-finance-bank-fd-rates/",
        "https://www.bankbazaar.com/fixed-deposit/ujjivan-small-finance-bank-fixed-deposit-rate.html",
        "https://www.paisabazaar.com/fixed-deposit/ujjivan-small-finance-bank-fd-rates/",
    ], "https://www.ujjivansfb.in/fixed-deposit-interest-rates"),
    ("Utkarsh SFB", [
        "https://www.policybazaar.com/fd-interest-rates/utkarsh-small-finance-bank-fd-rates/",
        "https://www.bankbazaar.com/fixed-deposit/utkarsh-small-finance-bank-fixed-deposit-rate.html",
        "https://www.paisabazaar.com/fixed-deposit/utkarsh-small-finance-bank-fd-rates/",
    ], "https://www.utkarsh.bank/fixed-deposit"),
    ("Unity SFB", [
        "https://www.policybazaar.com/fd-interest-rates/unity-small-finance-bank-fd-rates/",
        "https://www.bankbazaar.com/fixed-deposit/unity-small-finance-bank-fixed-deposit-rate.html",
    ], "https://theunitybank.com/fixed-deposit"),
    ("Suryoday SFB", [
        "https://www.policybazaar.com/fd-interest-rates/suryoday-small-finance-bank-fd-rates/",
        "https://www.bankbazaar.com/fixed-deposit/suryoday-small-finance-bank-fixed-deposit-rate.html",
    ], "https://www.suryodaybank.com/fixed-deposit"),
]


@st.cache_data(ttl=21600)
def fetch_all_fd():
    """Return list of (name, {tenor: rate}) tuples, fully automated.

    Stage 1 (parallel, HTTP): scrape every aggregator URL for each bank and
             MERGE the results, so a tenor missing on one site is filled by
             another. Thread-safe (no Streamlit cache calls, no Playwright).
    Stage 2 (sequential, MAIN thread): for any bank still missing tenors, open
             its aggregator page and then its own site in a headless browser to
             fill the gaps (handles JS-rendered tables that HTTP can't see)."""
    import concurrent.futures
    aggs_map = {n: aggs for n, aggs, _d in _FD_INSTITUTIONS}
    direct_map = {n: d for n, _a, d in _FD_INSTITUTIONS}
    by_name = {}

    # Stage 1 — all aggregator URLs, in parallel, merged per bank
    def _collect_http(item):
        name, aggs, _direct = item
        res = {}
        for u in aggs:
            _merge_tenors(res, _scrape_http(u))
            if len(res) == 5:
                break
        return name, res
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as ex:
        for name, res in ex.map(_collect_http, _FD_INSTITUTIONS):
            by_name[name] = res

    # Stage 2 — browser fill on the main thread, only for incomplete banks
    for name, aggs, direct in _FD_INSTITUTIONS:
        if len(by_name.get(name, {})) >= 5:
            continue
        for u in ([aggs[0]] if aggs else []) + ([direct] if direct else []):
            _merge_tenors(by_name[name], _scrape_browser(u))
            if len(by_name[name]) == 5:
                break

    return [(name, by_name[name]) for name, _a, _d in _FD_INSTITUTIONS]


@st.cache_data(ttl=3600)  # 1h — yields move daily; hourly is plenty
def bond_yield_browser(url):
    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        return None
    sels = ['[data-test="instrument-price-last"]',
            '[class*="instrument-price_last"]',
            '[class*="last_last"]']
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page(user_agent=_BROWSER_H["User-Agent"])
            page.goto(url, timeout=45000, wait_until="domcontentloaded")
            page.wait_for_timeout(5000)
            val = None
            for sel in sels:                       # try the precise element
                try:
                    el = page.query_selector(sel)
                    if el:
                        t = el.inner_text().strip().replace(",", "")
                        m = re.search(r"\d{1,2}\.\d{1,3}", t)
                        if m:
                            val = float(m.group()); break
                except Exception:
                    pass
            if val is None:                        # fallback: value before "Prev. Close"
                body = re.sub(r"\s+", " ", page.inner_text("body"))
                m = re.search(r"(\d{1,2}\.\d{1,3})\s+[\d.+%\-() ]*Prev\.?\s*Close", body, re.I)
                if not m:                          # last resort: first plausible number
                    m = re.search(r"\b(\d{1,2}\.\d{2,3})\b", body)
                if m:
                    val = float(m.group(1))
            browser.close()
        # sanity: government yields sit roughly 0–15%
        return val if (val is not None and 0 < val < 15) else None
    except Exception:
        return None


@st.cache_data(ttl=3600)  # 1h — yields move daily; hourly is plenty
def bond_yield_change_browser(url):
    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        return (None, None)
    sels = ['[data-test="instrument-price-last"]',
            '[class*="instrument-price_last"]',
            '[class*="last_last"]']
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page(user_agent=_BROWSER_H["User-Agent"])
            page.goto(url, timeout=45000, wait_until="domcontentloaded")
            page.wait_for_timeout(5000)
            val = None
            for sel in sels:                       # try the precise element
                try:
                    el = page.query_selector(sel)
                    if el:
                        m = re.search(r"\d{1,2}\.\d{1,3}",
                                      el.inner_text().strip().replace(",", ""))
                        if m:
                            val = float(m.group()); break
                except Exception:
                    pass
            body = re.sub(r"\s+", " ", page.inner_text("body"))
            if val is None:                        # fallback: value before "Prev. Close"
                m = re.search(r"(\d{1,2}\.\d{1,3})\s+[\d.+%\-() ]*Prev\.?\s*Close",
                              body, re.I)
                if m:
                    val = float(m.group(1))
            pc = re.search(r"Prev\.?\s*Close\s*(\d{1,2}\.\d{1,3})", body, re.I)
            prev_close = float(pc.group(1)) if pc else None
            browser.close()
        val = val if (val is not None and 0 < val < 15) else None
        change = None
        if val is not None and prev_close is not None:
            d = round(val - prev_close, 3)
            if abs(d) < 1.0:                       # ignore an implausible daily move
                change = d
        return (val, change)
    except Exception:
        return (None, None)


# ---------------------------------------------------------------------------
# FORMAT HELPERS
# ---------------------------------------------------------------------------
def chg(pct, bps=False, inverse=False):
    if pct is None:
        return '<span class="muted">-</span>'
    up = pct >= 0
    good = (not up) if inverse else up
    color = "var(--pos)" if good else "var(--neg)"
    arrow = "\u25B2" if up else "\u25BC"
    unit = " bps" if bps else "%"
    body = f"{pct:+.0f}" if bps else f"{abs(pct):.2f}"
    return f'<span style="color:{color}">{arrow} {body}{unit}</span>'


def num(x, d=2, prefix=""):
    if x is None:
        return '<span class="muted">-</span>'
    return prefix + f"{{:,.{d}f}}".format(x)


def gap(note="no free feed"):
    return f'<span class="muted">- <small>({note})</small></span>'


def _link(text, url):
    """Wrap text in a styled hyperlink (opens in new tab). Falls back to plain
    text if url is empty/None — safe to call unconditionally."""
    if not url:
        return str(text)
    return (f'<a href="{url}" target="_blank" rel="noopener" '
            f'style="color:inherit;text-decoration:underline dotted '
            f'rgba(91,140,255,.45);text-underline-offset:3px">{text}</a>')


# Source URLs for each ticker / item (used to make names clickable)
_YAHOO = "https://finance.yahoo.com/quote/"
_SRC = {
    # Equities
    "^NSEI": f"{_YAHOO}%5ENSEI", "^BSESN": f"{_YAHOO}%5EBSESN",
    "^GSPC": f"{_YAHOO}%5EGSPC", "^IXIC": f"{_YAHOO}%5EIXIC",
    "^DJI": f"{_YAHOO}%5EDJI", "^GDAXI": f"{_YAHOO}%5EGDAXI",
    "^FTSE": f"{_YAHOO}%5EFTSE", "000001.SS": f"{_YAHOO}000001.SS",
    "^HSI": f"{_YAHOO}%5EHSI", "^N225": f"{_YAHOO}%5EN225",
    # FX
    "INR=X": f"{_YAHOO}INR%3DX", "EURINR=X": f"{_YAHOO}EURINR%3DX",
    "GBPINR=X": f"{_YAHOO}GBPINR%3DX", "JPYINR=X": f"{_YAHOO}JPYINR%3DX",
    "CNY=X": f"{_YAHOO}CNY%3DX",
    # Commodities
    "GC=F": f"{_YAHOO}GC%3DF", "SI=F": f"{_YAHOO}SI%3DF",
    "BZ=F": f"{_YAHOO}BZ%3DF", "CL=F": f"{_YAHOO}CL%3DF",
    "NG=F": f"{_YAHOO}NG%3DF", "HG=F": f"{_YAHOO}HG%3DF",
    "ALI=F": f"{_YAHOO}ALI%3DF",
    # VIX
    "^INDIAVIX": f"{_YAHOO}%5EINDIAVIX", "^VIX": f"{_YAHOO}%5EVIX",
    # Alt assets
    "EMBASSY.BO": f"{_YAHOO}EMBASSY.BO", "MINDSPACE.BO": f"{_YAHOO}MINDSPACE.BO",
    "BIRET.BO": f"{_YAHOO}BIRET.BO", "NXST.BO": f"{_YAHOO}NXST.BO",
    "INDIGRID.BO": f"{_YAHOO}INDIGRID.BO", "PGINVIT.BO": f"{_YAHOO}PGINVIT.BO",
    # Crypto
    "BTC-USD": f"{_YAHOO}BTC-USD", "ETH-USD": f"{_YAHOO}ETH-USD",
}
# Rates sources (by country)
_RATE_SRC = {
    "India": "https://www.investing.com/rates-bonds/india-government-bonds",
    "US": "https://home.treasury.gov/resource-center/data-chart-center/interest-rates/daily-treasury-rates.csv/2026/all",
    "UK": "https://www.investing.com/rates-bonds/uk-government-bonds",
    "Germany": "https://data.ecb.europa.eu/data/data-categories/financial-markets-and-interest-rates/euro-area-yield-curves",
    "Japan": "https://www.mof.go.jp/english/policy/jgbs/reference/interest_rate/index.htm",
}
# Macro sources
# Source links point at the ORIGINAL government publisher wherever one exists,
# not at a third-party mirror — so every figure is one click from its primary
# release. FRED is used only for US series (where FRED *is* the aggregator of
# record for BLS/Fed data) and for India's 10Y yield.
_MACRO_SRC = {
    # MoSPI is the publisher of record for CPI (base 2024=100 from Feb 2026).
    "India CPI": "https://cpi.mospi.gov.in/",
    # DPIIT's Office of the Economic Adviser publishes WPI (base 2022-23).
    "India WPI": "https://eaindustry.nic.in/",
    # MoSPI National Accounts Statistics — GDP releases.
    "India GDP": "https://www.mospi.gov.in/data",
    "US CPI Index": "https://www.bls.gov/cpi/",
    "US Unemployment": "https://www.bls.gov/cps/",
    # RBI Weekly Statistical Supplement carries the forex reserves print.
    "India Forex": "https://www.rbi.org.in/Scripts/BS_ViewWSS.aspx",
}


def hcells(d, keys=("m3", "m6", "y1", "y5")):
    """3M / 6M / 1Y / 5Y change cells for a Yahoo record (gap where unavailable)."""
    if not d:
        return [gap() for _ in keys]
    return [chg(d[k]) if d.get(k) is not None else gap() for k in keys]


def _pcfmt(price, pct, prefix, dec):
    """Stacked cell: price on top, change below. Muted 'NA' when data missing."""
    if price is None or pct is None:
        return '<span class="muted">NA</span>'
    return (f'{prefix}{price:,.{dec}f}'
            f'<br><span style="font-size:.85em">{chg(pct)}</span>')


def pc(price, pct, prefix="", dec=2):
    return _pcfmt(price, pct, prefix, dec)


def pcells(d, labels, prefix="", dec=2):
    """Stacked price+change cells for 'd1' only; just the % change for all
    other horizons (1M, 3M, 6M, 1Y, 5Y, YTD). Keeps the table clean —
    current price is already in the 'Current' column."""
    if not d:
        return ['<span class="muted">NA</span>' for _ in labels]
    b = d.get("base", {})
    out = []
    for l in labels:
        pct = d.get(l)
        base_price = b.get(l)
        if l == "d1":
            # 1-day: show price + change (stacked)
            out.append(pc(base_price, pct, prefix, dec))
        else:
            # All other horizons: just the % change
            if pct is not None:
                out.append(chg(pct))
            else:
                out.append('<span class="muted">NA</span>')
    return out


def _inr(x):
    """Format a rupee amount: no decimals for large values, 2 for small."""
    if x is None:
        return None
    return f"\u20b9{x:,.0f}" if abs(x) >= 1000 else f"\u20b9{x:,.2f}"


def inr_pc(usd_now, usd_base, fx_now, fx_base, factor=1.0):
    """INR-terms % change only (no stacked price). Used for commodity change
    columns where the price is already shown in the Intl/Indian columns."""
    if usd_base is None or fx_now is None or fx_base is None:
        return '<span class="muted">NA</span>'
    inr_then = usd_base * fx_base * factor
    inr_now = usd_now * fx_now * factor
    pct = (inr_now / inr_then - 1) * 100
    return chg(pct)


# Fields that CAN be auto-pulled from FRED (free). suffix notes the cadence.
FRED_SERIES = {
    "india_10y":  ("INDIRLTLT01STM", "%", " (monthly)"),
    "japan_10y":  ("IRLTLT01JPM156N", "%", " (monthly)"),
    "uk_10y":     ("IRLTLT01GBM156N", "%", " (monthly)"),
    "germany_10y": ("IRLTLT01DEM156N", "%", " (monthly)"),
    "india_cpi":  ("INDCPIALLMINMEI", "", " (index, monthly)"),
    "india_call": ("IRSTCI01INM156N", "%", " (monthly)"),  # call money / interbank
}


def auto_val(field, note="no free feed"):
    """Manual typed value wins; else FRED auto; else a labelled gap."""
    v = str(manual.get(field, "")).strip()
    if v:
        return f'{v}'
    if field in FRED_SERIES and FRED_API_KEY:
        sid, unit, suffix = FRED_SERIES[field]
        val = fred(sid, FRED_API_KEY)
        if val is not None:
            return f'{val:.2f}{unit}'
    return gap(note)


_REPORT = []  # filled fresh each run; consumed by the download buttons


def _strip(c):
    """Plain text from an HTML cell, for the report files. Keeps <br> as a
    newline so price-over-change cells stay two lines in the export."""
    if c is None:
        return ""
    t = re.sub(r"<br\s*/?>", "\n", str(c))
    t = _html.unescape(_TAGS.sub("", t))
    return "\n".join(re.sub(r"\s+", " ", ln).strip() for ln in t.split("\n")).strip()


def report_section(title):
    """Start a report section AND render its heading with an anchor ID on the page."""
    _REPORT.append({"title": title, "tables": []})
    # Create a clean anchor ID from the title (e.g. "1 · Global equity markets" → "sec-1")
    sec_id = "sec-" + title.split("·")[0].strip().replace(" ", "").replace(".", "")
    st.markdown(f'<div id="{sec_id}"></div>', unsafe_allow_html=True)
    st.subheader(title)


def _sec_desc(text):
    """Render a one-line plain-English description under a section heading.
    Helps non-finance readers understand what they're looking at."""
    st.markdown(f'<div class="sec-desc">{text}</div>', unsafe_allow_html=True)


def render_table(headers, rows):
    th = "".join(f"<th>{h}</th>" for h in headers)
    trs = ""
    for r in rows:
        trs += "<tr>" + "".join(f"<td>{c}</td>" for c in r) + "</tr>"
    st.markdown(f'<div class="dash-wrap"><table class="dash">'
                f'<thead><tr>{th}</tr></thead>'
                f'<tbody>{trs}</tbody></table></div>', unsafe_allow_html=True)
    if _REPORT:  # record a clean copy for the downloadable report
        _REPORT[-1]["tables"].append(
            {"headers": list(headers),
             "rows": [[_strip(c) for c in r] for r in rows]})


def _pdf_safe(s):
    """reportlab core fonts are latin-1; swap symbols and drop the rest."""
    s = (str(s).replace("\u20b9", "Rs ").replace("\u25b2", "+")
         .replace("\u25bc", "-").replace("\u2013", "-").replace("\u00b7", "."))
    return s.encode("latin-1", "replace").decode("latin-1")


def _pdf_markup(s):
    """Like _pdf_safe but converts the dashboard's coloured HTML spans into
    reportlab <font color> markup, so up moves print green and down moves red.
    Hyperlinks (<a href>) are kept — reportlab renders them as clickable in PDF."""
    s = str(s)
    s = s.replace("color:var(--pos)", "C_POS").replace("color:var(--neg)", "C_NEG")
    s = re.sub(r'<span[^>]*C_POS[^>]*>', '<font color="#0a8f3c">', s)   # green
    s = re.sub(r'<span[^>]*C_NEG[^>]*>', '<font color="#c0392b">', s)   # red
    s = re.sub(r'<span class="muted"[^>]*>', '<font color="#9aa0a6">', s)
    s = re.sub(r'<span[^>]*>', '<font>', s)        # any other span -> plain font
    s = s.replace("</span>", "</font>")
    s = s.replace("<small>", "").replace("</small>", "")
    s = s.replace("<br>", "<br/>")
    # Simplify <a> tags for reportlab: keep only href, strip styling attributes
    s = re.sub(r'<a\s+href="([^"]*)"[^>]*>', r'<a href="\1" color="blue">', s)
    return _pdf_safe(s)


def chart_table():
    """Monthly 5-year history of the 10Y govt-bond yield for all 5 countries."""
    if not FRED_API_KEY:
        return None, {}
    series = {}
    for label, sid in (("India", "INDIRLTLT01STM"),
                       ("US", "IRLTLT01USM156N"),
                       ("UK", "IRLTLT01GBM156N"),
                       ("Germany", "IRLTLT01DEM156N"),
                       ("Japan", "IRLTLT01JPM156N")):
        h = fred_history(sid, FRED_API_KEY, years=5)
        if h is not None and len(h):
            series[label] = h.resample("MS").last()
    if not series:
        return None, {}
    df = pd.DataFrame(series).sort_index().tail(61)
    dates = [d.strftime("%b %Y") for d in df.index]
    cols = {c: [None if pd.isna(x) else round(float(x), 2) for x in df[c]]
            for c in df.columns}
    return dates, cols


def chart_png(dates, cols):
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except Exception:
        return None
    colors_map = {"India": "#1a73e8", "US": "#ff3d00", "UK": "#7c4dff",
                  "Germany": "#ff9100", "Japan": "#00bfa5"}
    widths_map = {"India": 2.0, "US": 2.2, "UK": 1.4, "Germany": 1.4, "Japan": 1.4}
    fig, ax = plt.subplots(figsize=(7.4, 3.2), dpi=150)
    x = list(range(len(dates)))
    for c in cols:
        ys = [v if v is not None else float("nan") for v in cols[c]]
        lw = widths_map.get(c, 1.5)
        color = colors_map.get(c, "#333")
        line, = ax.plot(x, ys, label=c, linewidth=lw, color=color)
        for i in range(0, len(x), 3):
            yi = ys[i] if i < len(ys) else float("nan")
            if yi == yi:
                ax.plot(i, yi, "o", color=line.get_color(), markersize=2.8)
                ax.annotate(f"{yi:.2f}", (i, yi), textcoords="offset points",
                            xytext=(0, 5), ha="center", fontsize=4.2,
                            color=line.get_color(), fontweight="bold")
    step = max(1, len(dates) // 8)
    ax.set_xticks(x[::step])
    ax.set_xticklabels(dates[::step], rotation=45, ha="right", fontsize=7)
    ax.set_title("10Y Government Bond Yield — 5-Year History", fontsize=9, fontweight="bold")
    ax.legend(fontsize=7, loc="upper left", framealpha=0.8)
    ax.grid(alpha=.2)
    fig.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format="png"); plt.close(fig)
    buf.seek(0)
    return buf


def market_summary(eq, fx, cm, gold):
    """Build a short, data-driven 'what matters today' summary from the live
    numbers. Returns a list of plain-text bullet lines. Everything is derived
    from the same figures shown in the tables — nothing hand-written."""
    lines = []

    def emv(tk):
        d = eq.get(tk)
        if d and d.get("d1") is not None:
            return d["price"], d["d1"]
        return None, None

    # India headline (Nifty + Sensex)
    npx, nmv = emv("^NSEI")
    _spx, smv = emv("^BSESN")
    if nmv is not None:
        seg = f"Nifty 50 at {npx:,.0f} ({nmv:+.2f}%)"
        if smv is not None:
            seg += f", Sensex {smv:+.2f}%"
        lines.append("India equities: " + seg + ".")

    # Leader / laggard on the day (index name + country in brackets)
    moves = [(name, region, eq[tk]["d1"]) for region, name, tk in EQUITIES
             if eq.get(tk) and eq[tk].get("d1") is not None]
    if moves:
        best = max(moves, key=lambda x: x[2])
        worst = min(moves, key=lambda x: x[2])
        lines.append(f"{best[0]} ({best[1]}) led ({best[2]:+.2f}%); "
                     f"{worst[0]} ({worst[1]}) lagged ({worst[2]:+.2f}%).")

    # Rupee
    u = fx.get("INR=X")
    if u and u.get("d1") is not None:
        lines.append(f"Rupee: USD/INR {u['price']:.2f} ({u['d1']:+.2f}% today).")

    # Gold (live Indian rate; day-move proxied by COMEX trend)
    g = (gold or {}).get("24k")
    gmv = cm.get("GC=F", {}).get("d1") if cm.get("GC=F") else None
    if g:
        seg = f"Gold 24K Rs {g * 10:,.0f}/10g"
        if gmv is not None:
            seg += f" ({gmv:+.2f}%)"
        lines.append(seg + ".")

    # What matters today: the largest absolute index move
    if moves:
        big = max(moves, key=lambda x: abs(x[2]))
        verb = "rose" if big[2] >= 0 else "fell"
        tone = ("a quiet session with small moves across the board"
                if abs(big[2]) < 0.5 else
                f"{big[0]} ({big[1]}) {verb} {abs(big[2]):.2f}%, the day's biggest move")
        lines.append("Biggest move: " + tone + ".")

    return lines


def build_excel(report, meta, summary=None):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Dashboard"
    bold = Font(bold=True)
    r = 1
    c = ws.cell(r, 1, "Daily Market Dashboard"); c.font = Font(bold=True, size=14); r += 1
    for m in meta:
        ws.cell(r, 1, m); r += 1
    r += 1
    if summary:
        c = ws.cell(r, 1, "Today's summary"); c.font = Font(bold=True, size=12); r += 1
        for s in summary:
            cell = ws.cell(r, 1, "• " + s)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
        r += 1
    for sec in report:
        c = ws.cell(r, 1, sec["title"]); c.font = Font(bold=True, size=12); r += 1
        for tbl in sec["tables"]:
            for ci, h in enumerate(tbl["headers"], 1):
                ws.cell(r, ci, h).font = bold
            r += 1
            for row in tbl["rows"]:
                for ci, val in enumerate(row, 1):
                    cell = ws.cell(r, ci, val)
                    if isinstance(val, str) and "\n" in val:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                r += 1
            r += 1
    for col in ws.columns:
        w = max((len(str(cell.value)) for cell in col if cell.value is not None),
                default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 60)

    # 5-year yield chart on its own sheet (native Excel line chart)
    try:
        dates, cols = chart_table()
        if dates and cols:
            from openpyxl.chart import LineChart, Reference
            cs = wb.create_sheet("5Y Yield Chart")
            cs.cell(1, 1, "Month")
            for j, c in enumerate(cols, start=2):
                cs.cell(1, j, c)
            for i, d in enumerate(dates, start=2):
                cs.cell(i, 1, d)
                for j, c in enumerate(cols, start=2):
                    cs.cell(i, j, cols[c][i - 2])
            ch = LineChart()
            ch.title = "10Y govt bond yield - 5-year trend (monthly)"
            ch.y_axis.title = "10Y yield (%)"; ch.x_axis.title = "Month"
            ch.height = 9; ch.width = 20
            data = Reference(cs, min_col=2, max_col=1 + len(cols),
                             min_row=1, max_row=1 + len(dates))
            cats = Reference(cs, min_col=1, min_row=2, max_row=1 + len(dates))
            ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
            cs.add_chart(ch, "G2")
    except Exception:
        pass

    bio = io.BytesIO(); wb.save(bio); return bio.getvalue()


def build_pdf(report, meta, summary=None):
    """Thin wrapper — the real builder lives in report_pdf.py, which the daily
    email (send_report.py) imports too, so both PDFs are always identical."""
    png = None
    try:
        dates, cols = chart_table()
        if dates and cols:
            png = chart_png(dates, cols)
    except Exception:
        pass
    return _report_pdf.build_pdf(report, meta, summary,
                                 chart_png=png, dashboard_url=DASHBOARD_URL)




# ---------------------------------------------------------------------------
# PAGE
# ---------------------------------------------------------------------------
st.set_page_config(page_title="Daily Market Dashboard", layout="wide")
st.markdown("""<style>
:root{--pos:#26d07c;--neg:#ff5c6c;--accent:#5b8cff;--muted:#8b98a9;
  --card-bg:linear-gradient(180deg,rgba(255,255,255,.045),rgba(255,255,255,.012));
  --card-bd:1px solid rgba(255,255,255,.08)}

/* Force dark mode on all devices, regardless of OS/browser light setting */
html, body, [data-testid="stAppViewContainer"], .stApp,
[data-testid="stHeader"], [data-testid="stSidebar"]{
  background-color:#0d1117 !important;
  color:#e6edf3 !important;
}
.mchg-lbl{font-size:.6em;color:var(--muted);font-weight:500;letter-spacing:.02em}

/* page rhythm */
.block-container{padding-top:2.2rem;max-width:1180px}

/* title */
h1{font-weight:800!important;letter-spacing:-.02em;
  background:linear-gradient(92deg,#eaf1ff 0%,#9db8ff 60%,#5b8cff 100%);
  -webkit-background-clip:text;background-clip:text;-webkit-text-fill-color:transparent}

/* section headings get an accent bar */
h3{position:relative;padding-left:14px!important;margin-top:.4rem!important;
  font-weight:700!important;letter-spacing:-.01em}
h3::before{content:"";position:absolute;left:0;top:.18em;bottom:.18em;width:4px;
  border-radius:3px;background:linear-gradient(180deg,var(--accent),#9db8ff)}


/* table card */
.dash-wrap{border:var(--card-bd);border-radius:14px;
  overflow-x:auto;-webkit-overflow-scrolling:touch;
  margin:6px 0 20px;background:var(--card-bg);
  box-shadow:0 1px 2px rgba(0,0,0,.25)}
table.dash{width:auto;min-width:100%;border-collapse:collapse;font-size:14px;margin:0}
table.dash thead th{text-align:right;color:var(--muted);font-weight:600;font-size:11px;
  letter-spacing:.05em;text-transform:uppercase;padding:11px 16px;white-space:nowrap;
  background:rgba(255,255,255,.045);border-bottom:1px solid rgba(255,255,255,.10)}
table.dash td{text-align:right;padding:10px 16px;white-space:nowrap;
  border-bottom:1px solid rgba(255,255,255,.05)}
table.dash tbody tr:nth-child(even){background:rgba(255,255,255,.022)}
table.dash tbody tr:hover{background:rgba(91,140,255,.10)}
table.dash tbody tr:last-child td{border-bottom:none}
table.dash th:first-child,table.dash td:first-child,
table.dash th:nth-child(2),table.dash td:nth-child(2){text-align:left}
table.dash td:first-child{font-weight:600}


/* management summary cards */
.mrow{display:flex;flex-wrap:wrap;gap:12px;margin:8px 0 6px}
.mcard{flex:1 1 150px;min-width:148px;padding:13px 16px;border-radius:13px;
  border:var(--card-bd);background:var(--card-bg);transition:transform .15s;cursor:pointer}
.mcard:hover{transform:translateY(-2px)}
.mname{color:var(--muted);font-size:11px;letter-spacing:.06em;text-transform:uppercase;
  margin-bottom:5px}
.mval{font-size:19px;font-weight:700;line-height:1.2}
.msrc{color:#6b7280;font-size:10px;margin-top:5px}
.bw{display:flex;gap:18px;flex-wrap:wrap;margin:10px 0 2px;color:var(--muted);font-size:13px}

/* what-matters-today summary box */
.summary-box{margin:14px 0 4px;padding:14px 18px;border-radius:13px;
  border:1px solid rgba(91,140,255,.22);
  background:linear-gradient(180deg,rgba(91,140,255,.08),rgba(91,140,255,.02))}
.summary-h{font-size:11px;letter-spacing:.07em;text-transform:uppercase;
  color:var(--accent);font-weight:700;margin-bottom:7px}
.summary-list{margin:0;padding-left:18px}
.summary-list li{font-size:14px;line-height:1.6;margin:1px 0}

.muted{color:var(--muted)}
.src{color:var(--muted);font-size:12px;margin:-4px 0 12px}

/* slim scrollbar */
.dash-wrap::-webkit-scrollbar{height:7px}
.dash-wrap::-webkit-scrollbar-thumb{background:rgba(255,255,255,.18);border-radius:8px}

/* ---- phones ---- */
@media (max-width:640px){
  .block-container{padding-top:1.2rem;padding-left:.6rem;padding-right:.6rem;max-width:100%}
  h1{font-size:1.55rem!important}
  table.dash{font-size:12.5px}
  table.dash thead th{padding:8px 10px;font-size:10px}
  table.dash td{padding:8px 10px}
  .mcard{flex:1 1 calc(50% - 12px);min-width:0;padding:11px 13px}
  .mval{font-size:17px}
  .mood-bar{flex-direction:column;text-align:center;gap:8px}
  .mood-icon{font-size:32px}
  .summary-box{padding:12px 14px}
  .summary-list li{font-size:13px}
  .bw{gap:12px;font-size:12px}
}
</style>""", unsafe_allow_html=True)

with st.sidebar:
    st.header("Settings")
    secs = st.select_slider("Auto-refresh every (sec)",
                            [30, 60, 120, 300], value=60)
    st.caption("Leave this tab open; it refreshes itself.")
    if st.button("Refresh now"):
        st.cache_data.clear(); st.rerun()

    with st.expander("FRED key (one-time, enables US/EU/India macro)"):
        st.caption("Free key from fredaccount.stlouisfed.org/apikeys. Saved on "
                   "this computer. This is a login, not market data.")
        k = st.text_input("FRED API key", value=str(manual.get("fred_key", "")))
        if st.button("Save key"):
            save_manual({"fred_key": k})
            st.success("Saved. Refreshing..."); st.rerun()

    with st.expander("India macro (CPI / WPI / GDP — update on release)"):
        st.caption("No free live API exists for these (MoSPI / DPIIT press releases). "
                   "Enter the latest print, e.g. '4.38% (Jun 2026)'. Leave blank to "
                   "use the built-in value, which also shows the change vs the "
                   "previous month.")
        c_in = st.text_input("India CPI inflation (YoY)",
                             value=str(manual.get("india_cpi", "")),
                             placeholder=f"{LATEST_INDIA_CPI[0]}% ({LATEST_INDIA_CPI[1]})")
        w_in = st.text_input("India WPI inflation (YoY)",
                             value=str(manual.get("india_wpi", "")),
                             placeholder=f"{LATEST_INDIA_WPI[0]}% ({LATEST_INDIA_WPI[1]})")
        g_in = st.text_input("India GDP growth",
                             value=str(manual.get("india_gdp", "")),
                             placeholder=f"{LATEST_INDIA_GDP[0]}% ({LATEST_INDIA_GDP[1]})")
        st.divider()
        st.caption("India G-Sec yields — FRED series discontinued. "
                   "Enter the yield as a number (e.g. 5.92). "
                   "Source: RBI / FBIL daily benchmark.")
        y1_in = st.text_input("India 1Y yield (%)",
                              value=str(manual.get("india_1y", "")),
                              placeholder=str(LATEST_INDIA_1Y))
        y2_in = st.text_input("India 2Y yield (%)",
                              value=str(manual.get("india_2y", "")),
                              placeholder=str(LATEST_INDIA_2Y))
        if st.button("Save India macro"):
            save_manual({"india_cpi": c_in.strip(), "india_gdp": g_in.strip(),
                         "india_wpi": w_in.strip(),
                         "india_1y": y1_in.strip(), "india_2y": y2_in.strip()})
            st.success("Saved. Refreshing..."); st.rerun()


if HAS_AUTOREFRESH:
    st_autorefresh(interval=secs*1000, key="auto")


# ---- Logo (top-right) + Title ----
try:
    _logo_file = os.path.join(os.path.dirname(__file__) or ".", "logo.png")
    _lcol1, _lcol2 = st.columns([3, 1])
    with _lcol1:
        st.title("Daily Market Dashboard")
    with _lcol2:
        if os.path.exists(_logo_file):
            st.image(_logo_file, width=160)
except Exception:
    st.title("Daily Market Dashboard")
st.caption(f"🕐 {dt.datetime.now():%a %d %b %Y  %H:%M:%S} {TZ}  ·  "
           f"Last updated {dt.datetime.now():%H:%M:%S} {TZ}")

# fetch
_REPORT.clear()  # rebuild the report fresh every run
eq = yahoo(EQUITIES); fx = yahoo(FX); vl = yahoo(VOL)
cm = yahoo([("c", t) for t in dict.fromkeys(r[1] for r in COMMODITIES)])
cry = yahoo(CRYPTO)
_gold = india_gold()      # live Indian retail 24K per gram (incl. duty/GST)
try:
    ty = treasury()
except Exception:
    ty = {"1Y": (None, None), "2Y": (None, None), "10Y": (None, None), "date": "-"}

# ---- Management summary ----



def mcard(name, value, change_html, src, anchor=""):
    a_open = f'<a href="#{anchor}" style="text-decoration:none;color:inherit">' if anchor else ''
    a_close = '</a>' if anchor else ''
    chg_label = f'{change_html} <span class="mchg-lbl">vs yest.</span>' if change_html else ''
    return (f'{a_open}<div class="mcard"><div class="mname">{name}</div>'
            f'<div class="mval">{value} {chg_label}</div></div>{a_close}')
def hl(name, tk, src, d, anchor=""):
    v = d.get(tk)
    return mcard(name, num(v['price']) if v else '-',
                 chg(v['d1']) if v else '', src, anchor)
cards = []
if eq.get("^NSEI"): cards.append(hl("Nifty 50", "^NSEI", "", eq, "sec-1"))
if eq.get("^GSPC"): cards.append(hl("S&P 500", "^GSPC", "", eq, "sec-1"))
if ty["10Y"][0] is not None:
    cards.append(mcard("US 10Y Yield", f"{ty['10Y'][0]:.2f}%",
                       chg(ty['10Y'][1]*100, bps=True, inverse=True), "", "sec-2"))
if cm.get("GC=F"): cards.append(hl("Gold", "GC=F", "", cm, "sec-4"))
if fx.get("INR=X"): cards.append(hl("USD/INR", "INR=X", "", fx, "sec-3"))
if cards:
    st.markdown('<div class="mrow">' + "".join(cards) + '</div>',
                unsafe_allow_html=True)
perf = [(n, eq[t]["d1"]) for _, n, t in EQUITIES if eq.get(t)]
if perf:
    b = max(perf, key=lambda x: x[1]); w = min(perf, key=lambda x: x[1])
    st.markdown(f'<div class="bw"><span>Best: <b>{b[0]}</b> {chg(b[1])}</span>'
                f'<span>Worst: <b>{w[0]}</b> {chg(w[1])}</span></div>',
                unsafe_allow_html=True)

# data-driven "What matters today" summary (same lines used in the PDF/Excel)
try:
    _summary = market_summary(eq, fx, cm, _gold)
except Exception:
    _summary = []
if _summary:
    _items = "".join(f"<li>{_html.escape(s)}</li>" for s in _summary)
    st.markdown(
        '<div class="summary-box"><div class="summary-h">Biggest move</div>'
        f'<ul class="summary-list">{_items}</ul></div>',
        unsafe_allow_html=True)
st.divider()

# ---- 1. Equities ----
report_section("1 · Global equity markets")
rows = []
for region, name, tk in EQUITIES:
    d = eq.get(tk)
    cur = num(d["price"]) if d else gap("fetch failed")
    linked_name = _link(name, _SRC.get(tk))
    rows.append([region, linked_name, cur] + pcells(d, ("d1", "m1", "m3", "m6", "y1", "ytd")))
render_table(["Region", "Index", "Current", "1D", "1M", "3M", "6M", "1Y", "YTD"], rows)



# ---- 2. Rates ----
report_section("2 · Interest rates & fixed income")

# Browser scrape (Investing.com) — works locally, blocked on Streamlit Cloud
YIELD_PAGES = {
    "india_1y":   "https://www.investing.com/rates-bonds/india-1-year-bond-yield",
    "india_2y":   "https://www.investing.com/rates-bonds/india-2-year-bond-yield",
    "uk_1y":      "https://www.investing.com/rates-bonds/uk-1-year-bond-yield",
    "uk_2y":      "https://www.investing.com/rates-bonds/uk-2-year-bond-yield",
    "japan_1y":   "https://www.investing.com/rates-bonds/japan-1-year-bond-yield",
    "japan_2y":   "https://www.investing.com/rates-bonds/japan-2-year-bond-yield",
}
YIELD_10Y_PAGES = {
    "india": "https://www.investing.com/rates-bonds/india-10-year-bond-yield",
    "uk":    "https://www.investing.com/rates-bonds/uk-10-year-bond-yield",
    "japan": "https://www.investing.com/rates-bonds/japan-10-year-bond-yield",
}
_yld = {}
_yld10 = {}
# Only attempt browser scraping if Playwright is installed (works locally,
# not on Streamlit Cloud). Cloud uses the API sources below instead.
try:
    from playwright.sync_api import sync_playwright as _pw_test
    for k, u in YIELD_PAGES.items():
        _yld[k] = bond_yield_browser(u)
    for k, u in YIELD_10Y_PAGES.items():
        _yld10[k] = bond_yield_change_browser(u)
except ImportError:
    pass

# Cloud-friendly sources (plain HTTP, no browser needed)
_jgb = japan_jgb()          # Japan MOF CSV
_ecb = ecb_yield()          # Germany ECB API
_boe = boe_yields()         # UK BoE CSV API
_ind_short = india_short_yields()  # India via FRED


def _yield_cell(country, tenor):
    """Pick the best available value for a country+tenor from multiple sources.
    Priority: cloud-friendly API → browser scrape → FRED monthly → gap."""
    # 1. Cloud-friendly dedicated fetchers
    if country == "japan":
        v = _jgb.get(tenor)
        if v is not None:
            return f"{v:.2f}%"
    if country == "germany":
        v = _ecb.get(tenor)
        if v is not None:
            return f"{v:.2f}%"
    if country == "uk" and tenor in ("1Y", "2Y"):
        v = _boe.get(tenor)
        if v is not None:
            return f"{v:.2f}%"
    if country == "india" and tenor in ("1Y", "2Y"):
        v = _ind_short.get(tenor)
        if v is not None:
            return f"{v:.2f}%"

    # 2. US: Treasury.gov daily curve, then FRED daily (must run BEFORE the
    #    generic 10Y fallback below, which has no 'us_10y' FRED_SERIES entry).
    if country == "us":
        tv = ty.get(tenor, (None, None))[0]
        if tv is not None:
            return f"{tv:.2f}%"
        fred_map = {"1Y": "DGS1", "2Y": "DGS2", "10Y": "DGS10"}
        fv, _ = fred_latest(fred_map.get(tenor, ""), FRED_API_KEY)
        if fv is not None:
            return f"{fv:.2f}%"
        return gap()

    # 3. Browser scrape (Investing.com — works locally)
    bkey = f"{country}_{tenor.lower()}"
    v = _yld.get(bkey)
    if v is not None:
        return f"{v:.2f}%"

    # 4. FRED monthly (10Y only for India/UK/Germany/Japan)
    if tenor == "10Y":
        return auto_val(f"{country}_10y")

    return gap()


render_table(["Country", "1Y", "2Y", "10Y"], [
    [_link("India", _RATE_SRC["India"]),
     _yield_cell("india", "1Y"), _yield_cell("india", "2Y"), _yield_cell("india", "10Y")],
    [_link("US", _RATE_SRC["US"]),
     _yield_cell("us", "1Y"), _yield_cell("us", "2Y"), _yield_cell("us", "10Y")],
    [_link("UK", _RATE_SRC["UK"]),
     _yield_cell("uk", "1Y"), _yield_cell("uk", "2Y"), _yield_cell("uk", "10Y")],
    [_link("Germany", _RATE_SRC["Germany"]),
     _yield_cell("germany", "1Y"), _yield_cell("germany", "2Y"), _yield_cell("germany", "10Y")],
    [_link("Japan", _RATE_SRC["Japan"]),
     _yield_cell("japan", "1Y"), _yield_cell("japan", "2Y"), _yield_cell("japan", "10Y")],
])
st.markdown('<div class="src">US = Treasury.gov + FRED; Germany = ECB euro-area AAA '
            'curve; Japan = MOF JGB curve; UK = Bank of England + FRED; '
            'India = FRED (short-term T-bill proxy for 1Y/2Y, 10Y direct).</div>',
            unsafe_allow_html=True)

# ---- 5-year trend: 10Y government bond yield, India / US / UK / Japan ----
# US comes back daily (DGS10); India/UK/Japan are FRED's monthly OECD series.
# Each line connects only its own real observations - nothing interpolated/faked.
TENY_SERIES = [
    ("US",      "DGS10",           "daily"),
    ("India",   "INDIRLTLT01STM",  "monthly"),
    ("UK",      "IRLTLT01GBM156N", "monthly"),
    ("Germany", "IRLTLT01DEM156N", "monthly"),
    ("Japan",   "IRLTLT01JPM156N", "monthly"),
]
_ORDER = ["India", "US", "UK", "Germany", "Japan"]
if FRED_API_KEY:
    _cut5 = pd.Timestamp(dt.date.today() - pd.Timedelta(days=365 * 5))
    _frames, _mark_frames = [], []
    for _label, _sid, _ in TENY_SERIES:
        _h = fred_history(_sid, FRED_API_KEY, years=5)
        if _h is not None and len(_h):
            _hh = _h[_h.index >= _cut5]
            _d = _hh.reset_index()
            _d.columns = ["date", "yield"]
            _d["country"] = _label
            _frames.append(_d)
            # one marker every 3 months (quarter starts), labelled with the value
            _q = _hh.resample("3MS").last().dropna()
            _m = _q.reset_index()
            _m.columns = ["date", "yield"]
            _m["country"] = _label
            _mark_frames.append(_m)
    if _frames:
        _long = pd.concat(_frames, ignore_index=True)
        _marks = pd.concat(_mark_frames, ignore_index=True)
        try:
            _x = _alt.X("date:T", title=None,
                             axis=_alt.Axis(format="%b %y", labelAngle=-45,
                                                 tickCount={"interval": "month",
                                                            "step": 3}))
            _color = _alt.Color("country:N", title=None, sort=_ORDER,
                                 scale=_alt.Scale(
                                     domain=["India", "US", "UK", "Germany", "Japan"],
                                     range=["#1a73e8", "#ff3d00", "#7c4dff", "#ff9100", "#00bfa5"]),
                                 legend=_alt.Legend(orient="bottom",
                                                    direction="horizontal",
                                                    columns=5,
                                                    labelFontSize=11,
                                                    symbolSize=80))
            _tip = ["country:N", "date:T", _alt.Tooltip("yield:Q", format=".2f")]
            _line = (_alt.Chart(_long).mark_line(strokeWidth=1.6)
                     .encode(x=_x,
                             y=_alt.Y("yield:Q", title=None,
                                           scale=_alt.Scale(zero=False)),
                             color=_color, tooltip=_tip))
            _dots = (_alt.Chart(_marks)
                     .mark_point(filled=True, size=26, opacity=1)
                     .encode(x=_x, y="yield:Q", color=_color, tooltip=_tip))
            _labels = (_alt.Chart(_marks)
                       .mark_text(dy=-9, fontSize=8, fontWeight="bold")
                       .encode(x=_x, y="yield:Q", color=_color,
                               text=_alt.Text("yield:Q", format=".2f")))
            _chart = (_line + _dots + _labels).properties(height=420)
            st.altair_chart(_chart, use_container_width=True)
        except Exception:
            # Fallback if Altair is unavailable: forward-fill monthly across days
            _wide = (_long.pivot(index="date", columns="country", values="yield")
                          .sort_index().ffill())
            st.line_chart(_wide, height=420)
    else:
        st.markdown('<div class="src">5-year chart: no data returned from FRED '
                    '(try "Refresh now", or check the key).</div>',
                    unsafe_allow_html=True)
else:
    st.markdown('<div class="src">Add a free FRED key in the sidebar to see the '
                '5-year yield chart.</div>', unsafe_allow_html=True)

with st.expander("Yield scrape details (verify the picked value)"):
    st.caption("India/UK/Japan 1Y, 2Y & 10Y come from Investing.com via headless "
               "browser. A gap means the page changed or the browser is blocked.")
    for k, u in YIELD_PAGES.items():
        v = _yld.get(k)
        st.markdown(f"**{k}** — {v:.2f}% from {u}" if v is not None
                    else f"**{k}** — no value read from {u}")
    for k, u in YIELD_10Y_PAGES.items():
        v, c = _yld10.get(k, (None, None))
        if v is not None:
            cs = (f", daily change {c:+.2f} ({c*100:+.0f} bps)" if c is not None
                  else ", daily change n/a (no prev-close read)")
            st.markdown(f"**{k}_10y** — {v:.2f}%{cs} from {u}")
        else:
            st.markdown(f"**{k}_10y** — no value read (FRED monthly used) from {u}")
ff = fred("FEDFUNDS", FRED_API_KEY)
rbi = bis_policy("IN")
st.markdown('<div class="sec-desc" style="margin-top:14px">Central bank policy rates — '
            'the rate that influences every loan, deposit and EMI. '
            'Changes shown in basis points (100 bps = 1%).</div>',
            unsafe_allow_html=True)

# Policy rates with the same 1M/3M/6M/1Y/5Y horizons as the other tables.
# FEDFUNDS = US Fed funds effective rate (FRED, monthly).
# INTDSRINM193N = India policy/discount rate (FRED, monthly) — used for history.
_ff_now, _ff_chg = policy_rate_changes("FEDFUNDS", FRED_API_KEY)
_rbi_now, _rbi_chg = policy_rate_changes("INTDSRINM193N", FRED_API_KEY)


def _bps(d, lbl):
    """Render a bps change cell (falling rate = green, in line with yields)."""
    v = d.get(lbl)
    if v is None:
        return '<span class="muted">NA</span>'
    return chg(v, bps=True, inverse=True)


_rbi_cell = (f"{_rbi_now:.2f}%" if _rbi_now is not None
             else (rbi if rbi else gap("no free feed")))
_ff_cell = (f"{_ff_now:.2f}%" if _ff_now is not None
            else (f"{ff:.2f}%" if ff is not None else gap("add FRED key")))

render_table(["Central bank", "Policy rate", "1M", "3M", "6M", "1Y"], [
    [_link("RBI Repo Rate", "https://www.rbi.org.in/Scripts/BS_ViewBulletin.aspx"),
     _rbi_cell] + [_bps(_rbi_chg, l) for l in ("m1", "m3", "m6", "y1")],
    [_link("US Fed Funds Rate", "https://fred.stlouisfed.org/series/FEDFUNDS"),
     _ff_cell] + [_bps(_ff_chg, l) for l in ("m1", "m3", "m6", "y1")],
])
st.markdown('<div class="src">Changes in basis points over each period. '
            'A falling policy rate shows green (cheaper borrowing).</div>',
            unsafe_allow_html=True)

# ---- 4. Currencies ----
report_section("3 · Currency & crypto markets")
_sec_desc("Exchange rates — how many rupees one dollar buys. A rising number means the rupee is weakening. Crypto shown in USD.")
rows = []
for name, tk in FX:
    d = fx.get(tk)
    cur = num(d["price"], 2) if d else gap()
    rows.append([_link(name, _SRC.get(tk)), cur] + pcells(d, ("d1", "m1", "m3", "m6", "y1", "y5"), "", 2))
# Bitcoin & Ethereum added to this table (in USD)
for name, tk in CRYPTO:
    d = cry.get(tk)
    cur = num(d["price"], 0, "$") if d else gap()
    rows.append([_link(name, _SRC.get(tk)), cur] + pcells(d, ("d1", "m1", "m3", "m6", "y1", "y5"), "$", 0))
render_table(["Pair / Asset", "Current", "1D", "1M", "3M", "6M", "1Y", "5Y"], rows)


# ---- 5. Commodities ----
report_section("4 · Commodities")
_inrfx = fx.get("INR=X")
_fxnow = _inrfx["price"] if _inrfx else None
_fxbase = _inrfx.get("base", {}) if _inrfx else {}
# _gold fetched once in the main fetch block above (live Indian retail 24K/g)

_INR_UNITS = {
    "Gold 24K": "INR/10g", "Silver": "INR/kg", "Brent Crude": "INR/bbl",
    "WTI Crude": "INR/bbl", "Natural Gas": "INR/MMBtu",
    "Copper": "INR/kg", "Aluminum": "INR/t",
}

rows = []
na = '<span class="muted">NA</span>'
for name, tk, inr_factor, intl_unit, intl_factor in COMMODITIES:
    d = cm.get(tk)
    if not d or _fxnow is None:
        rows.append([_link(name, _SRC.get(tk)), na, na] + [na] * 5)
        continue

    linked = _link(name, _SRC.get(tk))

    # International price (USD)
    intl_price = d["price"] * intl_factor
    intl_cell = f"${intl_price:,.2f} <small style='color:var(--muted)'>{intl_unit}</small>"

    # Indian price (INR)
    inu = _INR_UNITS.get(name, "")
    if name == "Gold 24K":
        g = _gold.get("24k")
        inr_val = _inr(g * 10) if g else _inr(d["price"] * _fxnow * inr_factor)
    else:
        inr_val = _inr(d["price"] * _fxnow * inr_factor)
    inr_cell = f"{inr_val} <small style='color:var(--muted)'>{inu}</small>"

    # Change columns (international USD %)
    b = d.get("base", {})
    cells = [inr_pc(d["price"], b.get(lbl), _fxnow, _fxbase.get(lbl), inr_factor)
             for lbl in ("d1", "m1", "m6", "y1", "y5")]
    rows.append([linked, intl_cell, inr_cell] + cells)

render_table(["Commodity", "Intl Price", "Indian Price",
              "1D", "1M", "6M", "1Y", "5Y"], rows)

st.markdown('<div class="src">Gold 24K Indian price = live retail rate (Goodreturns, '
            'incl. duty &amp; GST). All others = international spot × USD/INR.</div>',
            unsafe_allow_html=True)

# ---- 6. Alt assets ----
# ---- 5. Volatility ----
report_section("5 · Volatility & risk indicators")
_sec_desc("The 'fear gauge' — a higher VIX means bigger expected market swings. Below 15 is calm, above 25 is nervous.")
# Standard direction colouring so the arrow, colour and PDF sign all agree:
# up = green ▲ +, down = red ▼ −  (no inverse, which made a fall show green ▼).
_ivd = vl.get("^INDIAVIX")
_uvd = vl.get("^VIX")
render_table(["Indicator", "Current", "1D", "1M", "3M", "6M", "1Y"], [
    [_link("India VIX", _SRC.get("^INDIAVIX")),
     num(_ivd["price"]) if _ivd else gap()]
    + pcells(_ivd, ("d1", "m1", "m3", "m6", "y1"), "", 2),
    [_link("US VIX (CBOE)", _SRC.get("^VIX")),
     num(_uvd["price"]) if _uvd else gap()]
    + pcells(_uvd, ("d1", "m1", "m3", "m6", "y1"), "", 2),
])
st.markdown('<div class="src">A rising VIX means the market expects bigger swings. '
            'Changes are % moves in the index level over each period.</div>',
            unsafe_allow_html=True)


# ---- 6. Macro ----
report_section("6 · Macro indicators")
_sec_desc("Big-picture economic numbers with the period each figure covers. These move slowly (monthly/quarterly) but shape everything else.")
cpi = fred("CPIAUCSL", FRED_API_KEY); un = fred("UNRATE", FRED_API_KEY)
# Forex reserves: prefer FRED monthly (current, excl. gold); fall back to the
# World Bank annual (incl. gold, but lags by ~1-2 years) if no FRED key.
_resv, _resv_d = fred_latest("TRESEGINM052N", FRED_API_KEY)   # USD millions, monthly
if _resv is not None:
    _resv_cell = f"${_resv/1000:,.2f}B ({_resv_d}, excl. gold)"
else:
    fxr = worldbank("FI.RES.TOTL.CD")
    _resv_cell = (f"${fxr[0]/1e9:,.2f}B ({fxr[1]}, incl. gold)" if fxr
                  else gap("no free source"))
# India CPI / GDP / WPI have no free real-time API (MoSPI publishes by press
# release only, and FRED's OECD-MEI India series were discontinued). Show the
# latest published figure — sidebar override wins, else the seeded default.
_gdp_m = str(manual.get("india_gdp", "")).strip()
_wpi_m = str(manual.get("india_wpi", "")).strip()
_cpi_m = str(manual.get("india_cpi", "")).strip()


def _india_cells(const, manual_val):
    """(value cell, change cell) for an India macro row.
    const = (latest, period, previous). The change is the move in percentage
    points between the last two published prints — a real number, not a label."""
    if manual_val:
        return manual_val, '<span class="muted">manual entry</span>'
    latest, period, prev = const
    val = f"{latest}% <span class='muted'>({period})</span>"
    try:
        pp = float(latest) - float(prev)
        return val, chg(pp, bps=False)
    except Exception:
        return val, '<span class="muted">NA</span>'


_incpi_cell, _incpi_chg = _india_cells(LATEST_INDIA_CPI, _cpi_m)
_wpi_cell, _wpi_chg = _india_cells(LATEST_INDIA_WPI, _wpi_m)
_gdp_cell, _gdp_chg = _india_cells(LATEST_INDIA_GDP, _gdp_m)

# US series are live on FRED — show level, the observation month, and the
# year-on-year change so the number has context.
_uscpi_val, _uscpi_chg = _macro_cells("CPIAUCSL", FRED_API_KEY)
_usun_val, _usun_chg = _macro_cells("UNRATE", FRED_API_KEY, unit="%", as_change=True)
_resv_val, _resv_chg = _macro_cells("TRESEGINM052N", FRED_API_KEY)
if _resv_val.startswith("-") or "add FRED" in _resv_val:
    _resv_val, _resv_chg = _resv_cell, '<span class="muted">NA</span>'
else:
    _lvl, _yoy, _dt = macro_yoy("TRESEGINM052N", FRED_API_KEY)
    _resv_val = f"${_lvl/1000:,.2f}B <span class='muted'>({_dt}, excl. gold)</span>"

render_table(["Indicator", "Latest", "Change", "Period"], [
    [_link("India CPI inflation (YoY)", _MACRO_SRC["India CPI"]),
     _incpi_cell, _incpi_chg, "Monthly"],
    ["India WPI inflation (YoY)", _wpi_cell, _wpi_chg, "Monthly"],
    [_link("India GDP Growth", _MACRO_SRC["India GDP"]), _gdp_cell, _gdp_chg,
     "Quarterly / annual"],
    [_link("US CPI Index", _MACRO_SRC["US CPI Index"]), _uscpi_val, _uscpi_chg, "Monthly"],
    [_link("US Unemployment", _MACRO_SRC["US Unemployment"]), _usun_val, _usun_chg, "Monthly"],
    [_link("India Forex Reserves", _MACRO_SRC["India Forex"]), _resv_val, _resv_chg, "Monthly"],
])
st.markdown('<div class="src">Date in brackets is the period the figure covers. '
            'India CPI / WPI changes are the move in percentage points vs the '
            'previous month\'s print; GDP compares FY26 actual with the RBI\'s FY27 '
            'forecast. US CPI change is year-on-year inflation. India CPI, WPI and '
            'GDP come from MoSPI / DPIIT press releases (no free live API) — update '
            'them in the sidebar when a new print lands.</div>',
            unsafe_allow_html=True)

# ---- footer ----
st.divider()
st.caption("Verify figures at the source.")

# ---- downloadable daily report (built from everything rendered above) ----
_stamp = dt.datetime.now()
_meta = [f"Generated {_stamp:%a %d %b %Y  %H:%M:%S} {TZ}"]
_fname = f"market-dashboard-{_stamp:%Y-%m-%d}"
try:
    _summary = market_summary(eq, fx, cm, _gold)
except Exception:
    _summary = []
with st.sidebar:
    st.divider()
    st.markdown("**Download today's report**")
    if HAS_XLSX:
        try:
            st.download_button(
                "Excel (.xlsx)", data=build_excel(_REPORT, _meta, _summary),
                file_name=_fname + ".xlsx",
                mime="application/vnd.openxmlformats-officedocument."
                     "spreadsheetml.sheet")
        except Exception as e:
            st.caption(f"Excel error: {e}")
    else:
        st.caption("Excel needs the openpyxl library "
                   "(run: pip install openpyxl).")
    if HAS_PDF:
        try:
            st.download_button(
                "PDF (.pdf)", data=build_pdf(_REPORT, _meta, _summary),
                file_name=_fname + ".pdf", mime="application/pdf")
        except Exception as e:
            st.caption(f"PDF error: {e}")
    else:
        st.caption("PDF needs the reportlab library "
                   "(run: pip install reportlab).")
    st.caption("Click a button to download a dated snapshot of today's numbers.")
